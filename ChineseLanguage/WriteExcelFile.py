
import os
import xlwt
import xlsxwriter 
import openpyxl as op
from xlwt import Workbook

class write_excel_file:
    filename = ''
    payload = ''
    url = ''

    def __init__(self, url, filename, payload):
        self.url = url 
        self.filename = filename
        self.payload = payload

    def write_contexts(self, attrs, htmls, scripts, urls):
        # creating the file for the first time 
        if not os.path.exists(self.filename):
            workbook = xlsxwriter.Workbook(self.filename)
            worksheet = workbook.add_worksheet("data")
            workbook.close()
            wb = op.load_workbook(self.filename, False)
            ws = wb['data']
            ws.append(["webpage", "payload", "# attr ", "# html ", "# script", "# url ",
                        "attr ", "html ", "script ", "url "    ])
            wb.save(self.filename)
            wb.close()

        print('\n\t\t --------- Writing to EXCEL ------------')
        wb = op.load_workbook(self.filename, False)
        ws = wb['data']
        ws.append([ self.url,   self.payload,
                    len(attrs), len(htmls), len(scripts), len(urls),
                    '',     '',     '',     '' ] )

        if not attrs == 'None':
            for attr in attrs:
                ws.cell(row=ws.max_row, column=7).value += str(attr) + ' , '
        else:
            ws.cell(row=ws.max_row, column=7).value += str(attrs)

        if not htmls == 'None':
            for html in htmls:  
                ws.cell(row=ws.max_row, column=8).value += str(html) + ' , '
        else:
            ws.cell(row=ws.max_row, column=8).value += str(htmls)

        if not scripts == 'None':
            for script in scripts:
                ws.cell(row=ws.max_row, column=9).value += str(script) + ' , '
        else:
            ws.cell(row=ws.max_row, column=9).value += str(scripts)

        if not urls == 'None':
            for url in urls:
                ws.cell(row=ws.max_row, column=10).value += str(url) + ' , '
        else:
            ws.cell(row=ws.max_row, column=10).value += str(urls)
            
        

       

        wb.save(self.filename)
        wb.close()


print('{WriteExcelFile}')